from flask import current_app
import openpyxl
import os
from copy import copy

import openpyxl.styles

def stylize_sheet(file_name):
    file_path = os.path.join(current_app.config.get('UPLOAD_FOLDER', 'somefile.txt'), file_name)
    file = openpyxl.load_workbook(file_path) 
    try:
        file.remove(file.worksheets[1])
    except:
        pass
    calc = file.worksheets[0]
    
    try:
       calc.unmerge_cells('A1:H1')
       calc['H1'].value = None
       calc.merge_cells('A1:G1')
    except:
       pass
    calc.delete_cols(8)
    
    
    for row in range(2, calc.max_row):
        source_cell = calc.cell(row=row, column=6)
        destination_cell = calc.cell(row=row, column=7)
        destination_cell.value = 'Original Price RI (USD)' if row == 2 else source_cell.value
        destination_cell.font = copy(source_cell.font) # type: ignore
        if calc.max_row - 1 == row:
            destination_cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='medium', color="000000"),
                                                             right=openpyxl.styles.Side(style='medium', color="000000"))
        else:
            destination_cell.border = openpyxl.styles.Border(right=openpyxl.styles.Side(style='medium', color="000000"))
    file.save(file_path)
    
def change_values(data):
    file_path = os.path.join(current_app.config.get('UPLOAD_FOLDER', 'somefile.txt'), data['uploadedFile'])
    file = openpyxl.load_workbook(file_path)
    calc = file.worksheets[0]
    calc[calc.max_row - 1][5].value = f"=SUM(F3:F{calc.max_row - 2})/2"
    calc[calc.max_row - 1][6].value = f"=SUM(G3:G{calc.max_row - 2})/2"
    for flavor in data['flavors']:
        calc[flavor['row'] - 1][5].value = f"=SUM(F{flavor['row']}:F{item_subindex(flavor['row'], calc)})"
        calc[flavor['row'] - 1][6].value = f"=SUM(G{flavor['row']}:G{item_subindex(flavor['row'], calc)})"
        if flavor['windows']:
            costs = get_cost_windows(flavor['name'])
            calc[flavor['row']][5].value = costs[0] # type: ignore
            calc[flavor['row']][6].value = costs[1] # type: ignore
            calc[flavor['row'] + 1][1].value = 'Windows Server' # type: ignore
        else: 
            calc[flavor['row']][6].value = get_cost_linux(flavor['name'])
    file.save(file_path)

def get_cost_windows(flavor):
    windows_prices = openpyxl.load_workbook(current_app.config.get('WINDOWS_PRICES_DIR', 'somefile.txt')).worksheets[1]
    try:
        for row in windows_prices.iter_rows(min_row=2, values_only=True):
            if row[11] == flavor:
                return (float(row[17]), float(row[18])) # type: ignore
    except:
        return (None, None)
        
def get_cost_linux(flavor):
    linux_prices = openpyxl.load_workbook(current_app.config.get('LINUX_PRICES_DIR', 'somefile.txt')).worksheets[0]
    try:
        for row in linux_prices.iter_rows(min_row=2, values_only=True):
            if row[0] == flavor:
                return float(row[1]) # type: ignore
    except:
        return None
    
def item_subindex(index, calc):
    cell_format = calc.cell(row=index, column=6).fill.start_color.rgb
    subindex = index
    print(cell_format)
    
    while True:
        subindex += 1
        current_cell = calc.cell(row=subindex, column=6).fill.start_color.rgb
        print(current_cell)
        if current_cell != cell_format:
            subindex -= 1
            return subindex
    
